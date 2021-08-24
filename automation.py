import os
import openpyxl as O
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

script_dir = os.path.dirname(__file__)

excel = {
    'file': script_dir + '/LP + Campanhas Atualizado.xlsx',
    'new': script_dir + '/result.xlsx',
    'worksheet': "Copy of LP + Campanhas Atualiza",
}

xpath = {
    'input': '//*[@id="page-speed-insights"]/div[1]/form/div/input', 
    'button':'//*[@id="page-speed-insights"]/div[1]/form/div/div/div',
}
result_class = 'lh-gauge__percentage'

url = "https://developers.google.com/speed/pagespeed/insights/"

max_wait=120

#setting headless mode
chrome_options = Options()
chrome_options.add_argument('--headless')
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument('--disable-dev-shm-usage')
driver = webdriver.Chrome(options=chrome_options)

#function to get scores
def get_score(page):
    driver.get(url)
    #searching page
    driver.find_element(By.XPATH, xpath['input']).send_keys(page)
    driver.find_element(By.XPATH, xpath['button']).click()
    #waiting for score elements and getting them
    wait=WebDriverWait(driver,max_wait)
    mob = wait.until(EC.presence_of_element_located((By.CLASS_NAME, result_class)))
    desk = driver.find_elements_by_class_name(result_class)[1]
    #extracting scores
    mob = mob.get_attribute('innerHTML')
    desk = desk.get_attribute('innerHTML')

    return (mob,desk)

#funcao para input de scores na planilha
def input_score(r):
    page = sheet.cell(r,1).value
    mob_cell = sheet.cell(r,col-1)
    desk_cell = sheet.cell(r,col)
    try:
        mob,desk = get_score(page)
        print(mob,desk)
        mob_cell.value=int(mob)
        desk_cell.value=int(desk)
    except:
        print('erro')
        mob_cell.value=desk_cell.value='erro'
    finally:
        wb.save(excel['new'])


#opening excel and extracting main info
wb=O.load_workbook(excel['file'])
sheet = wb[excel['worksheet']]
col = sheet.max_column
row_num = sheet.max_row
while True:
    if sheet.cell(row_num, 1).value not in [None,'']:
        print(sheet.cell(row_num, 1).value)
        break
    else:
        row_num -= 1

#making prep modifications
sheet.insert_cols(col-1,2)
sheet.cell(3,col-1).value = "Mobile"
sheet.cell(3,col).value = "Desktop"

# inputing each score
for r in range(4,row_num+1):
    input_score(r)

print('######################## END #######################')